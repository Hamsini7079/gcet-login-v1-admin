"""
=============================================================
  Geethanjali College of Engineering and Technology
  Flask Login Portal — VERSION 1
  app.py

  STORAGE:
    - SQLite database  → gcet_sessions.db   (reliable, multi-user)
    - Excel file       → gcet_sessions.xlsx (teachers can open easily)

  HOW TO RUN LOCALLY:
    1. pip install flask openpyxl
    2. python app.py
    3. Open http://localhost:5000

  HOW TO DEPLOY ON RENDER.COM:
    - Build Command : pip install -r requirements.txt
    - Start Command : python app.py
    - Add environment variable:
        KEY   → SECRET_KEY
        VALUE → any long random string
=============================================================
"""

from flask import Flask, render_template, request, jsonify
from werkzeug.utils import secure_filename
from datetime import datetime
import os, re, secrets, sqlite3, threading

# openpyxl is used to read and write Excel files
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ── SECRET KEY ─────────────────────────────────────────────
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

# ── SECURITY HEADERS ───────────────────────────────────────
@app.after_request
def add_security_headers(response):
    response.headers["X-Content-Type-Options"]  = "nosniff"
    response.headers["X-Frame-Options"]          = "DENY"
    response.headers["X-XSS-Protection"]         = "1; mode=block"
    response.headers["Referrer-Policy"]           = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"]        = "geolocation=(), microphone=(), camera=()"
    response.headers["Content-Security-Policy"]   = (
        "default-src 'self'; "
        "style-src  'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src   'self' https://fonts.gstatic.com; "
        "script-src 'self' 'unsafe-inline'; "
        "img-src    'self' data: blob:; "
        "connect-src 'self';"
    )
    return response

# ── FILE UPLOAD ────────────────────────────────────────────
UPLOAD_FOLDER   = os.path.join("static", "uploads")
ALLOWED_EXTS    = {"png", "jpg", "jpeg", "gif", "webp"}
MAX_UPLOAD_SIZE = 2 * 1024 * 1024
app.config["UPLOAD_FOLDER"]      = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_SIZE
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ── FILE PATHS ─────────────────────────────────────────────
DB_FILE    = "gcet_sessions.db"      # SQLite database file
EXCEL_FILE = "gcet_sessions.xlsx"    # Excel file for teachers

# ── WHITELIST VALIDATION ───────────────────────────────────
ALLOWED_ROLES    = {"student", "faculty"}
ALLOWED_BRANCHES = {"CSE", "ECE", "EEE", "Mechanical", "Civil"}
ALLOWED_SPECS    = {"AIML", "Cyber Security", "Data Science", ""}
ALLOWED_DEPTS    = {"CSE", "ECE", "EEE", "Mechanical", "Civil"}

# ── THREAD LOCK ────────────────────────────────────────────
# Prevents two people writing to Excel at the exact same time
excel_lock = threading.Lock()

# ── INPUT LIMITS ───────────────────────────────────────────
MAX_NAME = 80
MAX_ID   = 30


# ──────────────────────────────────────────────────────────
# SANITIZE INPUT
# ──────────────────────────────────────────────────────────
def sanitize(text, max_len=80):
    if not isinstance(text, str):
        return ""
    text = re.sub(r"<[^>]*>", "", text)
    text = re.sub(r"[<>\"';()&+\\]", "", text)
    return text.strip()[:max_len]


def now_str():
    return datetime.now().strftime("%d %b %Y, %I:%M:%S %p")


def calc_duration(entry_str, exit_str):
    fmt = "%d %b %Y, %I:%M:%S %p"
    try:
        delta  = datetime.strptime(exit_str, fmt) - datetime.strptime(entry_str, fmt)
        total  = int(delta.total_seconds())
        h, rem = divmod(total, 3600)
        m, s   = divmod(rem, 60)
        if h:   return f"{h}h {m}m {s}s"
        elif m: return f"{m}m {s}s"
        else:   return f"{s}s"
    except Exception:
        return "—"


# ══════════════════════════════════════════════════════════
# SQLITE DATABASE FUNCTIONS
# SQLite is built into Python — no extra install needed.
# It stores data in a single file: gcet_sessions.db
# ══════════════════════════════════════════════════════════

def get_db():
    """Open a connection to the SQLite database."""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row   # lets us access columns by name
    return conn


def init_db():
    """
    Create the sessions table if it doesn't exist yet.
    This runs once when the app starts.
    Columns:
      id           - unique auto number
      role         - student or faculty
      name         - full name
      branch       - branch or department
      specialization - CSE specialization (if any)
      roll_or_id   - roll number or faculty ID
      system_no    - system number (students only)
      entry_time   - when they logged in
      exit_time    - when they logged out
      duration     - how long they were logged in
    """
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            role           TEXT NOT NULL,
            name           TEXT NOT NULL,
            branch         TEXT,
            specialization TEXT,
            roll_or_id     TEXT,
            system_no      TEXT,
            entry_time     TEXT NOT NULL,
            exit_time      TEXT,
            duration       TEXT
        )
    """)
    conn.commit()
    conn.close()
    print("[DB] Database ready ✅")


def db_insert_session(record):
    """Insert a new login record into the database."""
    conn = get_db()
    cursor = conn.execute("""
        INSERT INTO sessions
            (role, name, branch, specialization, roll_or_id, system_no, entry_time)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        record["role"],
        record["name"],
        record.get("branch") or record.get("department", ""),
        record.get("specialization", ""),
        record.get("roll") or record.get("facultyId", ""),
        record.get("system", ""),
        record["entryTime"]
    ))
    row_id = cursor.lastrowid   # get the ID of the new row
    conn.commit()
    conn.close()
    return row_id


def db_update_exit(row_id, exit_time, duration):
    """Update exit time and duration when someone logs out."""
    conn = get_db()
    conn.execute("""
        UPDATE sessions
        SET exit_time = ?, duration = ?
        WHERE id = ?
    """, (exit_time, duration, row_id))
    conn.commit()
    conn.close()


def db_get_active():
    """Return the session with no exit time (currently logged in)."""
    conn = get_db()
    row = conn.execute("""
        SELECT * FROM sessions
        WHERE exit_time IS NULL
        ORDER BY id DESC LIMIT 1
    """).fetchone()
    conn.close()
    return dict(row) if row else None


def db_close_all_open(exit_time):
    """Close all open sessions (safety measure on new login)."""
    conn = get_db()
    rows = conn.execute("""
        SELECT id, entry_time FROM sessions WHERE exit_time IS NULL
    """).fetchall()
    for row in rows:
        dur = calc_duration(row["entry_time"], exit_time)
        conn.execute("""
            UPDATE sessions SET exit_time=?, duration=? WHERE id=?
        """, (exit_time, dur, row["id"]))
    conn.commit()
    conn.close()


def db_get_history(limit=15):
    """Get last N completed sessions for history display."""
    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM sessions
        WHERE exit_time IS NOT NULL
        ORDER BY id DESC LIMIT ?
    """, (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def db_clear_history():
    """Delete all completed sessions from the database."""
    conn = get_db()
    conn.execute("DELETE FROM sessions WHERE exit_time IS NOT NULL")
    conn.commit()
    conn.close()


# ══════════════════════════════════════════════════════════
# EXCEL FUNCTIONS
# Uses openpyxl to read and write .xlsx files.
# Excel is used so teachers can open and view sessions easily.
# ══════════════════════════════════════════════════════════

# Excel column headers
EXCEL_HEADERS = [
    "S.No", "Role", "Name", "Branch / Dept",
    "Specialization", "Roll No / Faculty ID",
    "System No", "Entry Time", "Exit Time", "Duration"
]

# Header row background color (teal)
HEADER_FILL  = PatternFill("solid", fgColor="00B4A6")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
# Alternating row colors
ROW_FILL_1   = PatternFill("solid", fgColor="EEF7F7")   # light teal
ROW_FILL_2   = PatternFill("solid", fgColor="FFFFFF")   # white
# Border style for cells
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)


def excel_init():
    """
    Create the Excel file with headers if it doesn't exist yet.
    Also styles the header row nicely.
    """
    if os.path.exists(EXCEL_FILE):
        return   # already exists, don't recreate

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GCET Sessions"

    # Write header row
    for col, header in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER

    # Set column widths
    col_widths = [6, 10, 22, 20, 18, 22, 12, 24, 24, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 22
    wb.save(EXCEL_FILE)
    print("[Excel] Excel file created ✅")


def excel_add_row(record):
    """
    Add a new login row to the Excel file.
    Uses a thread lock so two people can't write at the same time.
    """
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active

            # Count existing rows (excluding header)
            next_row = ws.max_row + 1
            s_no     = next_row - 1   # serial number

            row_data = [
                s_no,
                "Student" if record["role"] == "student" else "Faculty",
                record["name"],
                record.get("branch") or record.get("department", ""),
                record.get("specialization", "") or "—",
                record.get("roll") or record.get("facultyId", ""),
                record.get("system", "") or "—",
                record["entryTime"],
                record.get("exitTime") or "—",
                record.get("duration") or "—"
            ]

            # Alternate row colors for readability
            fill = ROW_FILL_1 if s_no % 2 == 0 else ROW_FILL_2

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=next_row, column=col, value=value)
                cell.fill      = fill
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

            wb.save(EXCEL_FILE)
        except Exception as e:
            print(f"[Excel ERROR] Could not add row: {e}")


def excel_update_exit(name, entry_time, exit_time, duration):
    """
    Find the matching row in Excel by name + entry time
    and update its exit time and duration columns.
    """
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                # Column 3 = Name, Column 8 = Entry Time
                if (row[2].value == name and
                    row[7].value == entry_time):
                    row[8].value = exit_time    # Column 9 = Exit Time
                    row[9].value = duration     # Column 10 = Duration
                    break

            wb.save(EXCEL_FILE)
        except Exception as e:
            print(f"[Excel ERROR] Could not update exit: {e}")


def excel_clear_history():
    """
    Remove all data rows from Excel except the header.
    Re-creates the file with just the header.
    """
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            # Delete all rows after the header (row 1)
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row)
            wb.save(EXCEL_FILE)
        except Exception as e:
            print(f"[Excel ERROR] Could not clear: {e}")


# ──────────────────────────────────────────────────────────
# HELPER — convert a DB row dict to the format the
# HTML template expects (same keys as before)
# ──────────────────────────────────────────────────────────
def db_row_to_session(row):
    """Convert a database row into a session dict for the template."""
    if not row:
        return None
    s = {
        "id":        row["id"],
        "role":      row["role"],
        "name":      row["name"],
        "entryTime": row["entry_time"],
        "exitTime":  row["exit_time"],
        "duration":  row["duration"],
    }
    if row["role"] == "student":
        s["branch"]         = row["branch"]
        s["specialization"] = row["specialization"] or ""
        s["roll"]           = row["roll_or_id"]
        s["system"]         = row["system_no"]
    else:
        s["department"] = row["branch"]
        s["facultyId"]  = row["roll_or_id"]
    return s


# ──────────────────────────────────────────────────────────
# ROUTES
# ──────────────────────────────────────────────────────────

@app.route("/")
def index():
    active_row = db_get_active()
    active     = db_row_to_session(active_row)
    hist_rows  = db_get_history(15)
    history    = [db_row_to_session(r) for r in hist_rows]

    logo_path = None
    for ext in ALLOWED_EXTS:
        candidate = os.path.join(UPLOAD_FOLDER, f"college_logo.{ext}")
        if os.path.exists(candidate):
            logo_path = f"/static/uploads/college_logo.{ext}"
            break

    return render_template("index.html",
                           active=active,
                           history=history,
                           logo_path=logo_path)


@app.route("/login", methods=["POST"])
def login():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"success": False, "message": "Invalid request."}), 400

    role = data.get("role", "")
    if role not in ALLOWED_ROLES:
        return jsonify({"success": False, "message": "Invalid role."}), 400

    name = sanitize(data.get("name", ""), MAX_NAME)
    if not name:
        return jsonify({"success": False, "message": "Name is required."}), 400

    record = {
        "role":      role,
        "name":      name,
        "entryTime": now_str(),
        "exitTime":  None,
        "duration":  None,
    }

    if role == "student":
        branch = data.get("branch", "")
        spec   = data.get("specialization", "")
        roll   = sanitize(data.get("roll",   ""), MAX_ID)
        system = sanitize(data.get("system", ""), MAX_ID)

        if branch not in ALLOWED_BRANCHES:
            return jsonify({"success": False, "message": "Invalid branch."}), 400
        if branch == "CSE" and spec not in ALLOWED_SPECS:
            return jsonify({"success": False, "message": "Invalid specialization."}), 400
        if not roll or not system:
            return jsonify({"success": False, "message": "Roll and System number are required."}), 400

        record.update({"branch": branch,
                       "specialization": spec if branch == "CSE" else "",
                       "roll": roll, "system": system})
    else:
        dept       = data.get("department", "")
        faculty_id = sanitize(data.get("facultyId", ""), MAX_ID)

        if dept not in ALLOWED_DEPTS:
            return jsonify({"success": False, "message": "Invalid department."}), 400
        if not faculty_id:
            return jsonify({"success": False, "message": "Faculty ID is required."}), 400

        record.update({"department": dept, "facultyId": faculty_id})

    # ── Close any open sessions first ──
    db_close_all_open(record["entryTime"])

    # ── Save to DATABASE ──
    row_id = db_insert_session(record)
    record["id"] = row_id

    # ── Save to EXCEL ──
    excel_add_row(record)

    return jsonify({"success": True, "session": record})


@app.route("/logout", methods=["POST"])
def logout():
    active_row = db_get_active()
    if not active_row:
        return jsonify({"success": False, "message": "No active session found."}), 404

    exit_time = now_str()
    duration  = calc_duration(active_row["entry_time"], exit_time)

    # ── Update DATABASE ──
    db_update_exit(active_row["id"], exit_time, duration)

    # ── Update EXCEL ──
    excel_update_exit(active_row["name"], active_row["entry_time"],
                      exit_time, duration)

    session = db_row_to_session(active_row)
    session["exitTime"] = exit_time
    session["duration"] = duration

    return jsonify({"success": True, "session": session})


@app.route("/clear_history", methods=["POST"])
def clear_history():
    db_clear_history()
    excel_clear_history()
    return jsonify({"success": True})


@app.route("/upload_logo", methods=["POST"])
def upload_logo():
    if "logo" not in request.files:
        return jsonify({"success": False, "message": "No file received."}), 400

    file = request.files["logo"]
    if not file or file.filename == "":
        return jsonify({"success": False, "message": "No file selected."}), 400

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_EXTS:
        return jsonify({"success": False,
                        "message": f"Allowed types: {', '.join(ALLOWED_EXTS)}"}), 400

    file.seek(0, 2); size = file.tell(); file.seek(0)
    if size > MAX_UPLOAD_SIZE:
        return jsonify({"success": False, "message": "Max file size is 2 MB."}), 400

    for old_ext in ALLOWED_EXTS:
        old = os.path.join(UPLOAD_FOLDER, f"college_logo.{old_ext}")
        if os.path.exists(old):
            os.remove(old)

    safe_name = f"college_logo.{ext}"
    file.save(os.path.join(app.config["UPLOAD_FOLDER"], safe_name))
    return jsonify({"success": True, "path": f"/static/uploads/{safe_name}"})


# ── Error handlers ─────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Page not found."}), 404

@app.errorhandler(413)
def too_large(e):
    return jsonify({"success": False, "message": "File too large. Max 2 MB."}), 413

@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": "Server error. Please try again."}), 500


# ──────────────────────────────────────────────────────────
# STARTUP
# ──────────────────────────────────────────────────────────


# ══════════════════════════════════════════════════════════
# ADMIN DASHBOARD
# ══════════════════════════════════════════════════════════

from flask import session as flask_session, redirect, url_for, send_file

# ── ADMIN PASSWORD ─────────────────────────────────────────
# Change this to your own password before publishing!
ADMIN_PASSWORD = "gcet@admin2024"


def is_admin():
    """Check if current user is logged in as admin."""
    return flask_session.get("is_admin") == True


@app.route("/admin", methods=["GET", "POST"])
def admin_login():
    """Admin login page."""
    error = None
    if request.method == "POST":
        password = request.form.get("password", "")
        if password == ADMIN_PASSWORD:
            flask_session["is_admin"] = True
            return redirect(url_for("admin_dashboard"))
        else:
            error = "Wrong password. Please try again."
    if is_admin():
        return redirect(url_for("admin_dashboard"))
    return render_template("admin_login.html", error=error)


@app.route("/admin/dashboard")
def admin_dashboard():
    """Admin dashboard — shows all session records."""
    if not is_admin():
        return redirect(url_for("admin_login"))
    conn     = get_db()
    all_rows = conn.execute("SELECT * FROM sessions ORDER BY id DESC").fetchall()
    today    = datetime.now().strftime("%d %b %Y")
    today_count  = conn.execute("SELECT COUNT(*) FROM sessions WHERE entry_time LIKE ?", (today+"%",)).fetchone()[0]
    active_count = conn.execute("SELECT COUNT(*) FROM sessions WHERE exit_time IS NULL").fetchone()[0]
    total_count  = conn.execute("SELECT COUNT(*) FROM sessions").fetchone()[0]
    conn.close()
    sessions = [db_row_to_session(dict(r)) for r in all_rows]
    return render_template("admin_dashboard.html",
                           sessions=sessions,
                           today_count=today_count,
                           active_count=active_count,
                           total_count=total_count,
                           today=today)


@app.route("/admin/download_excel")
def download_excel():
    """
    Generate Excel fresh from database every time download is clicked.
    This works correctly on Render.com because it reads from
    the database (which is reliable) instead of a saved file.
    """
    if not is_admin():
        return redirect(url_for("admin_login"))

    import io

    # Read ALL sessions from database
    conn = get_db()
    rows = conn.execute("SELECT * FROM sessions ORDER BY id ASC").fetchall()
    conn.close()

    # Create a fresh Excel workbook in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GCET Sessions"

    # Write header row
    for col, header in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER

    # Set column widths
    col_widths = [6, 10, 22, 20, 18, 22, 12, 24, 24, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22

    # Write all session rows
    for idx, row in enumerate(rows, start=1):
        s = dict(row)
        fill = ROW_FILL_1 if idx % 2 == 0 else ROW_FILL_2
        row_data = [
            idx,
            "Student" if s["role"] == "student" else "Faculty",
            s["name"],
            s.get("branch") or "—",
            s.get("specialization") or "—",
            s.get("roll_or_id") or "—",
            s.get("system_no") or "—",
            s.get("entry_time") or "—",
            s.get("exit_time") or "—",
            s.get("duration") or "—",
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=idx+1, column=col, value=value)
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save to memory buffer instead of a file
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="GCET_Sessions.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/admin/clear", methods=["POST"])
def admin_clear():
    """Clear all history."""
    if not is_admin():
        return redirect(url_for("admin_login"))
    db_clear_history()
    excel_clear_history()
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/logout")
def admin_logout():
    """Logout from admin."""
    flask_session.pop("is_admin", None)
    return redirect(url_for("admin_login"))


# ──────────────────────────────────────────────────────────
# RUN
# ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    excel_init()
    print("\n🎓 GCET Login Portal — Version 1")
    print("   SQLite DB  -> gcet_sessions.db")
    print("   Excel File -> gcet_sessions.xlsx")
    print("   Admin Page -> http://localhost:5000/admin")
    print("   Open in browser: http://localhost:5000\n")
    app.run(debug=False, host="0.0.0.0", port=5000)
